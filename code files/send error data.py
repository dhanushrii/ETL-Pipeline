import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import pandas as pd
import pyodbc
import paramiko
import io
from openpyxl.utils import get_column_letter
import os
from dotenv import load_dotenv

load_dotenv()

# ================= DB CONNECTION =================

conn = pyodbc.connect(
    f"DRIVER={{{os.getenv('DB_DRIVER')}}};"
    f"SERVER={os.getenv('DB_SERVER')};"
    f"DATABASE={os.getenv('DB_NAME')};"
    f"Trusted_Connection=yes;"
)


# ================= SFTP CONFIG =================

host = os.getenv("SFTP_HOST")
port = int(os.getenv("SFTP_PORT", 22))
sftp_user = os.getenv("SFTP_USER")
sftp_pass = os.getenv("SFTP_PASS")
remote_path = os.getenv("SFTP_PATH")

# ================= CREATE EXCEL =================

def create_excel(group_df):
    import io
    from openpyxl.utils import get_column_letter

    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        group_df.to_excel(writer, index=False, sheet_name='Errors')
        worksheet = writer.sheets['Errors']

        for i, col in enumerate(group_df.columns, 1):
            try:
                max_length = max(
                    group_df[col].fillna('').astype(str).map(lambda x: len(x)).max(),
                    len(col)
                )
                worksheet.column_dimensions[get_column_letter(i)].width = max_length + 2
            except:
                worksheet.column_dimensions[get_column_letter(i)].width = len(col) + 2

    buffer.seek(0)
    return buffer.read()

# ================= EMAIL FUNCTION =================

def send_email_with_attachment(to_email, subject, body, file_bytes, filename):
    sender = os.getenv("EMAIL_USER")
    password = os.getenv("EMAIL_PASS")


    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = to_email
    msg["Subject"] = subject

    msg.attach(MIMEText(body, "plain"))

    part = MIMEBase("application", "octet-stream")
    part.set_payload(file_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={filename}")

    msg.attach(part)

    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(sender, password)
            server.send_message(msg)
        print("✅ Email sent")

    except Exception as e:
        print("❌ Email failed:", e)

# ================= SFTP FUNCTION =================

def send_sftp(content, filename):
    try:
        transport = paramiko.Transport((host, port))
        transport.connect(username=sftp_user, password=sftp_pass)
        sftp = paramiko.SFTPClient.from_transport(transport)

        with sftp.file(f"{remote_path}/{filename}", "w") as f:
            f.write(content)

        sftp.close()
        transport.close()
        print("✅ SFTP uploaded:", filename)

    except Exception as e:
        print("❌ SFTP failed:", e)

# ================= FETCH DATA =================

query = """
SELECT 
    e.*,
    p.[mode of communication]
FROM error_table e
LEFT JOIN PartnerCode p
ON TRY_CAST(e.partner_code AS BIGINT) = p.[partner code]
"""

df = pd.read_sql(query, conn)

# ================= HANDLE NULL MODE =================

df["mode of communication"] = df["mode of communication"].fillna("UNKNOWN")

# ================= GROUPING =================

groups = df.groupby("mode of communication")

# ================= PROCESS =================

for mode, group in groups:

    mode = str(mode).strip().upper()

    print(f"\nProcessing mode: {mode} | Records: {len(group)}")

    # ---------- EMAIL ----------
    if mode == "EMAIL":

        excel_file = create_excel(group)

        send_email_with_attachment(
            to_email="22n215@psgtech.ac.in",
            subject="Error Report",
            body="Attached is the error report.",
            file_bytes=excel_file,
            filename="error_report.xlsx"
        )

    # ---------- SFTP ----------
    elif mode == "SFTP":

        text_data = group.to_csv(index=False)
        filename = "error_batch.txt"

        send_sftp(text_data, filename)

    # ---------- API ----------
    elif mode == "API":
        print("🔹 API call needed")
        print(group.head())

    # ---------- UNKNOWN ----------
    else:
        print("⚠ Unknown mode")
        print(group.head())

# ================= CLOSE =================

conn.close()